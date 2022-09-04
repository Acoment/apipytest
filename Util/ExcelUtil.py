# _*_ coding:UTF-8 _*_
# @time：2022/7/10  16:30
# @Author: 皓雪
# @file: ExcelUtil.py   
# @Software: PyCharm


from Conf.setting import DATA_PATH
import openpyxl

def get_row_values(file_name,sheet_name,row_no):
    """
    得到某一行的值，以列表形式返回
    :param file_name:
    :param sheet_name:
    :param row_no:
    :return:
    """

    file_path = f'{DATA_PATH}/{file_name}'
    wookbok = openpyxl.load_workbook(file_path)
    sheet = wookbok[sheet_name]
    data = []
    columns = sheet.max_column
    for i in range(1,columns+1):
        data.append(sheet.cell(row =row_no,column=i).value)
    return data

def get_test_data(file_name,sheet_name):
    """
    得到测试数目。返回[[],[],..]
    :param file_name:
    :param sheet_name:
    :return:
    """
    file_path = f'{DATA_PATH}/{file_name}'
    wookbok = openpyxl.load_workbook(file_path)
    sheet = wookbok[sheet_name]
    row_data = []
    row_nums = sheet.max_row
    columns = sheet.max_column
    for i in range(2,row_nums+1):
        sing_row_data = []
        for j in range(1,columns+1):
            sing_row_data.append(sheet.cell(row = i , column =j).value)
        row_data.append(sing_row_data)
    return row_data

if __name__ == '__main__':
    print(get_row_values('users.xlsx','login',1))
    print(get_row_values('users.xlsx', 'login'))
